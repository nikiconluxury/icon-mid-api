from fastapi import FastAPI, BackgroundTasks
import asyncio , os, requests, threading
from icon_image_lib.utility import process_row  # Assuming this is correctly implemented
import openpyxl
from openpyxl import load_workbook
# from icon_image_lib.utility import save_and_upload_file, send_completion_email
import uvicorn,imghdr
from PIL import Image
import uuid


app = FastAPI()

@app.post("/process-image-batch/")
async def process_payload(payload: dict):
    rows = payload.get('rowData', [])
    provided_file_path = payload['filePath']
    send_to_email = payload['sendToEmail']
    preferred_image_method = payload['preferredImageMethod']
    semaphore = asyncio.Semaphore(150)  # Limit to 150 concurrent tasks

    # Create a temporary directory to save downloaded images
    temp_dir = 'temp_images/'
    temp_dir = os.path.join(os.getcwd(), temp_dir + str(generate_unique_id_for_path()[:8]))
    os.makedirs(temp_dir, exist_ok=True)

    async def process_with_semaphore(row):
        async with semaphore:
            return await process_row(row)

    tasks = [process_with_semaphore(row) for row in rows]
    results = await asyncio.gather(*tasks)

    # Placeholder for aggregating results and updating the Excel file
    # You need to replace this with your actual logic for updating the Excel
    clean_results = prepare_data_for_download(results)
    print(clean_results)
    download_all_images(clean_results, temp_dir)
    #insert images
    
    
        #download each url once all downloaded
        
        #return saved image folder path
        #saved_img_folder_path = '' # placeholder for the saved image folder path
        #return saved_img_folder_path
    # Assume update_excel_file is a function you will implement to update the Excel file
    #update_excel_file(file_path, saved_img_folder_path)
    
    new_excel_filepath = ''# placeholder for the updated file path
    # Optionally upload the updated file and get the file URL
    uploaded_file_url = await save_and_upload_file(new_excel_filepath)  # Implement this function

    # Send completion email (implement this function according to your needs)
    await send_completion_email(send_to_email, "Batch Processing Completed", uploaded_file_url)

    return {"message": "Processing completed.", "results": results}

def update_excel_file(excel_file_path, results):
    # Implement your logic to open the Excel file, update it with results, and save it
    pass

async def save_and_upload_file(excel_file_path):
    # Implement your logic to save and upload the updated Excel file to a cloud storage
    # Return the URL to the uploaded file
    return "https://example.com/updated_excel.xlsx"

async def send_completion_email(to_email, subject, file_url):
    # Implement your logic to send an email with the completion status and file URL
    pass
def prepare_data_for_download(results):
    # Filter results to ensure each has a non-empty and non-None 'absoluteRowIndex' and 'url'
    filtered_results = [
        (result['absoluteRowIndex'], result['result']['url'])
        for result in results
        if result.get('absoluteRowIndex') is not None and 
           result.get('result') and 
           result['result'].get('url') not in [None, '', '[]']
    ]
    return filtered_results

def download_all_images(data, save_path):
        s = requests.Session()
        cookies = dict(BCPermissionLevel='PERSONAL')
        threads = []
        for item in data:
            # Extract the URL from the string (assuming only one URL is present in the string)
            image_url = item[1].strip("[]'\"")  # Remove the brackets and quotes
            input_sku = item[0]  # Grab the input SKU
            temp = threading.Thread(target=imageDownload, args=(str(image_url), str(input_sku), save_path,s,cookies))
            temp.start()
            threads.append(temp)
            
        for thread in threads:
            thread.join()
def generate_unique_id_for_path():
    return str(uuid.uuid4())
def imageDownload( url, imageName, newpath,s,cookies):
        timeout = 30
        headers = {
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3",
            "accept-encoding": "gzip, deflate, br",
            "accept-language": "en-US,en;q=0.9",
            "cookie": "Cookie: Something",
            "upgrade-insecure-requests": "1",
            "user-agent": "Mozilla/5.0"
        }
        if ('/' in imageName) or ('\\' in imageName):
            return False
        if ".webp" in str(url):
            with open(newpath + "/" + imageName + '.webp', 'wb') as handle:
                response = s.get(url, headers=headers, timeout=timeout, stream=True)
                if not response.ok:
                    print(response)
                for block in response.iter_content(1024):
                    if not block:
                        break

                    handle.write(block)
            im = Image.open(newpath + "/" + imageName + '.webp').convert("RGB")
            im.save(newpath + "/" + imageName + '.png', 'png')
        elif ".mpo" in str(url):
            print("broken image")
        # ! logging.debug("broken image")
        else:
            try:
                with open(newpath + "/" + imageName + '.png', 'wb') as handle:
                    response = s.get(url, headers=headers, timeout=timeout, stream=True)
                    if not response.ok:
                        print(response)
                    for block in response.iter_content(1024):
                        if not block:
                            break

                        handle.write(block)

                ######WEBP DETECTION            
                whatType = imghdr.what(newpath + "/" + imageName + '.png')
                print(whatType)
                ######WEBP DETECTION 
                if whatType == 'webp':
                    # ! logging.debug("Webp image detected")
                    with open(newpath + "/" + imageName + '.webp', 'wb') as handle:
                        response = s.get(url, headers={"User-Agent": "Mozilla/5.0"}, cookies=cookies,
                                                stream=True)
                        if not response.ok:
                            print(response)
                        for block in response.iter_content(1024):
                            if not block:
                                break

                            handle.write(block)
                    im = Image.open(newpath + "/" + imageName + '.webp').convert("RGB")
                    im.save(newpath + "/" + imageName + '.png', 'png')

            except Exception as exc:
                # !  logging.error("IMAGE DOWNLOAD ERROR:")
                print(exc)
                # !  logging.error(exc)
                print(imageName)
                # !  logging.error(imageName)
                print(url)
            # !  logging.error(url)

        return True
def write_excel_image(self, output):
        wb = load_workbook(self.filepath)
        ws = wb.active

        path = r"C:\Users\User\Documents\front-msrp\msrp_project\msrp_app\Images"
        # self.generate_unique_id_for_path()
        # path = path +"\\" + str(self.generate_unique_id_for_path())
        download_all_images(output, path)

        for image in output:
            # image is a tuple with the first element being the row number and the second element being the image link
            row_number = int(image[0]) - 1
            image_path = f"{path}/{image[1]}.png"
            if self.preferred_image_method == "O":
                    if os.path.exists(image_path):
                        image_g = self.verify_png_image_single(image_path)
                        if image_g:
                            resize = self.resize_image(image_path)
                            if resize:
                                img = openpyxl.drawing.image.Image(image_path)
                                img.anchor = "A" + str(row_number)
                                ws.add_image(img)
            elif self.preferred_image_method == "A":
                if not row_number in rows_with_images:
                    if os.path.exists(image_path):
                        image_g = self.verify_png_image_single(image_path)
                        if image_g:
                            resize = self.resize_image(image_path)
                            if resize:
                                img = openpyxl.drawing.image.Image(image_path)
                                img.anchor = "A" + str(row_number)
                                ws.add_image(img)
            elif self.preferred_image_method == "MNC":
                if os.path.exists(image_path):
                        image_g = self.verify_png_image_single(image_path)
                        if image_g:
                            resize = self.resize_image(image_path)
                            if resize:
                                img = openpyxl.drawing.image.Image(image_path)
                                img.anchor = "B" + str(row_number)
                                ws.add_image(img)
                    # ws.add_image(img, ws.cell(row=row_number, column=self.find_col_index("Images")).coordinate)

        self.finalize_changes(wb)

if __name__ == "__main__":
    # For local development with auto-reload
    uvicorn.run("main:app", port=8085, host='0.0.0.0', reload=True)
    # For production
    # uvicorn.run(app, port=8085, host='0.0.0.0')
