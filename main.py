from fastapi import FastAPI, BackgroundTasks
import asyncio, os,threading,uuid,requests,openpyxl,uvicorn,shutil,mimetypes,time
from icon_image_lib.utility import process_row  # Assuming this is correctly implemented
from openpyxl import load_workbook
from PIL import Image as IMG2
from PIL import UnidentifiedImageError
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
import datetime
import boto3
import logging
from io import BytesIO
from openpyxl.utils import get_column_letter
from icon_image_lib.google_parser import get_original_images as GP
from requests.adapters import HTTPAdapter
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib3.util.retry import Retry
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition,Personalization,Cc,To
from base64 import b64encode
from concurrent.futures import ThreadPoolExecutor, as_completed
import aiohttp
from aiohttp import ClientTimeout
from aiohttp_retry import RetryClient, ExponentialRetry
#logging.basicConfig(level=logging.INFO)
#logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
# Example usage
logger.info("Informational message")
logger.error("Error message")
from sqlalchemy import create_engine

import pyodbc
from dotenv import load_dotenv
import pandas as pd
load_dotenv()
import base64,zlib
from threading import Thread
def get_spaces_client():
    logger.info("Creating spaces client")
    session = boto3.session.Session()
    client = session.client('s3',
                            region_name='nyc3',
                            endpoint_url=os.getenv('SPACES_ENDPOINT'),
                            aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
                            aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY'))
    logger.info("Spaces client created successfully")
    return client

def upload_file_to_space(file_src, save_as, is_public, content_type, meta=None):
    spaces_client = get_spaces_client()
    space_name = 'iconluxurygroup-s3'  # Your space name
    print('Content Type')
    print(content_type)
    if not content_type:
        content_type_guess = mimetypes.guess_type(file_src)[0]
        if not content_type_guess:
            raise Exception("Content type could not be guessed. Please specify it directly.")
        content_type = content_type_guess

    extra_args = {'ACL': 'public-read' if is_public else 'private', 'ContentType': content_type}
    if meta:
        extra_args['Metadata'] = meta

    spaces_client.upload_file(
        Filename=file_src,
        Bucket=space_name,
        Key=save_as,
        ExtraArgs=extra_args
    )
    print(f"File uploaded successfully to {space_name}/{save_as}")
    # Generate and return the public URL if the file is public
    if is_public:
        upload_url = f"{str(os.getenv('SPACES_ENDPOINT'))}/{space_name}/{save_as}"
        print(f"Public URL: {upload_url}")
        return upload_url



def send_email(to_emails, subject, download_url, excel_file_path,execution_time,message="Total Rows:\nFilename:\nBatch ID:\nLocation:\nUploaded File:"):
    # Encode the URL if necessary (example shown, adjust as needed)
    # from urllib.parse import quote
    # download_url = quote(download_url, safe='')
    execution_time_timedelta = datetime.timedelta(seconds=execution_time)



    message_with_breaks = message.replace("\n", "<br>")

    html_content = f"""
<html>
<body>
<div class="container">
    <p>Your file is ready for download.</p>
    <p>Total Elapsed Time: {str(execution_time_timedelta)}</p>
    <p>Message details:<br>{message_with_breaks}</p>
    <a href="{download_url}" class="download-button">Download File</a>
    <p>CMS:v1</p>
</div>
</body>
</html>
"""




#     html_content = f"""
# <html>
# <body>
# <div class="container">
#     <p>Your file is ready for download.</p>
#     <p>Total Elapsed Time: {str(execution_time_timedelta)}</p>
#     <a href="{download_url}" class="download-button">Download File</a>
# </div>
# </body>
# </html>
# """
    message = Mail(
        from_email='distrotool@iconluxurygroup.com',
        subject=subject,
        html_content=html_content
    )
    # Read and encode the Excel file
    with open(excel_file_path, 'rb') as f:
        excel_data = f.read()
    encoded_excel_data = b64encode(excel_data).decode()

    attachment = Attachment(
        FileContent(encoded_excel_data),
        FileName(excel_file_path.split('/')[-1]),
        FileType('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
        Disposition('attachment')
    )
    message.attachment = attachment
    
    cc_recipient = 'notifications@popovtech.com'
    personalization = Personalization()
    personalization.add_cc(Cc(cc_recipient))
    personalization.add_to(To(to_emails))
    message.add_personalization(personalization)
    try:
        sg = SendGridAPIClient(os.environ.get('SENDGRID_API_KEY'))
        response = sg.send(message)
        print(response.status_code)
        print(response.body)
        print(response.headers)
    except Exception as e:
        print(e)
        #raise
        
def send_message_email(to_emails, subject,message):
    message_with_breaks = message.replace("\n", "<br>")

    html_content = f"""
<html>
<body>
<div class="container">
    <!-- Use the modified message with <br> for line breaks -->
    <p>Message details:<br>{message_with_breaks}</p>
    <p>CMS:v1</p>
</div>
</body>
</html>
"""
    message = Mail(
        from_email='distrotool@iconluxurygroup.com',
        subject=subject,
        html_content=html_content
    )
    
    cc_recipient = 'notifications@popovtech.com'
    personalization = Personalization()
    personalization.add_cc(Cc(cc_recipient))
    personalization.add_to(To(to_emails))
    message.add_personalization(personalization)
    try:
        sg = SendGridAPIClient(os.environ.get('SENDGRID_API_KEY'))
        response = sg.send(message)
        print(response.status_code)
        print(response.body)
        print(response.headers)
    except Exception as e:
        print(e)
        #raise
        
        
async def create_temp_dirs(unique_id):
    loop = asyncio.get_running_loop()  # Get the current loop directly
    base_dir = os.path.join(os.getcwd(), 'temp_files')
    temp_images_dir = os.path.join(base_dir, 'images', unique_id)
    temp_excel_dir = os.path.join(base_dir, 'excel', unique_id)

    await loop.run_in_executor(None, lambda: os.makedirs(temp_images_dir, exist_ok=True))
    await loop.run_in_executor(None, lambda: os.makedirs(temp_excel_dir, exist_ok=True))

    return temp_images_dir, temp_excel_dir

async def cleanup_temp_dirs(directories):
    loop = asyncio.get_running_loop()  # Get the current loop directly
    for dir_path in directories:
        await loop.run_in_executor(None, lambda dp=dir_path: shutil.rmtree(dp, ignore_errors=True))

pwd_value = str(os.environ.get('MSSQLS_PWD'))
pwd_str =f"Pwd={pwd_value};"
global conn
conn = "DRIVER={ODBC Driver 17 for SQL Server};Server=35.172.243.170;Database=luxurymarket_p4;Uid=luxurysitescraper;" + pwd_str
global engine
engine = create_engine("mssql+pyodbc:///?odbc_connect=%s" % conn)
app = FastAPI()
def insert_file_db (file_name,file_source):
    connection = pyodbc.connect(conn)
    cursor = connection.cursor()
    insert_query = "INSERT INTO utb_ImageScraperFiles (FileName, FileLocationUrl) OUTPUT INSERTED.Id VALUES (?, ?)"
    values = (file_name, file_source)

    cursor.execute(insert_query, values)

    file_id = cursor.fetchval()

    connection.commit()
    cursor.close()
    connection.close()

    return file_id
def get_records_to_search(file_id,engine):
    sql_query = f"Select EntryID, ProductModel as SearchString from utb_ImageScraperRecords where FileID = {file_id} and Step1 is null UNION ALL Select EntryID, ProductModel + ' '  + ProductBrand as SearchString from utb_ImageScraperRecords where FileID = {file_id} and Step1 is null Order by 1"
    print(sql_query)
    df = pd.read_sql_query(sql_query, con=engine)
    return df

def load_payload_db(rows, file_id):
    # Create DataFrame from list of dictionaries (rows)
    df = pd.DataFrame(rows)

    # Rename columns
    df = df.rename(columns={
        'absoluteRowIndex': 'ExcelRowID',  # Renaming 'index' to 'ExcelRowID'
        'searchValue': 'ProductModel',  # Renaming 'SKU' to 'ProductModel'
        'brandValue': 'ProductBrand',  # Renaming 'Brand' to 'ProductBrand'
        'colorValue': 'ProductColor',
        'CategoryValue': 'ProductCategory'
    })

    # Insert new column 'FileID' at the beginning with all values set to file_id
    df.insert(0, 'FileID', file_id)
    df = df.drop(columns=['imageValue'], axis=1)
    print(df)
    # Load DataFrame into SQL database
    df.to_sql(name='utb_ImageScraperRecords', con=engine, index=False, if_exists='append')

    return df
def get_endpoint():
     connection = pyodbc.connect(conn)
     cursor = connection.cursor()
     sql_query = "Select top 1 EndpointURL from utb_Endpoints where EndpointIsBlocked = 0 Order by NewID() "
     cursor.execute(sql_query)
     endpoint_url = cursor.fetchone()
     connection.commit()
     cursor.close()
     connection.close()
     if endpoint_url:
         (endpoint,) = endpoint_url
         print(endpoint_url)
         print(datetime.datetime.now())
     else:
         print("No EndpointURL")
         endpoint = "No EndpointURL"
     return endpoint
def remove_endpoint(endpoint):
    connection = pyodbc.connect(conn)
    cursor = connection.cursor()
    sql_query = f"Update utb_Endpoints set EndpointIsBlocked = 1 where  EndpointURL  = '{endpoint}'"
    cursor.execute(sql_query)
    connection.commit()
    cursor.close()
    connection.close()
def unpack_content(encoded_content):
    if encoded_content:
        compressed_content = base64.b64decode(encoded_content)
        original_content = zlib.decompress(compressed_content)
        # with open('text.html', 'w') as file:
        #     file.write(str(original_content))
        return original_content # Return as binary data
    return None
def process_search_row(search_string,endpoint,entry_id):
        search_url = f"{endpoint}?query={search_string}"
        print(search_url)

        try:
            response = requests.get(search_url, timeout=60)
            print(response.status_code)
            if response.status_code != 200 or response.json().get('body') is None:
                print('trying again 1')
                remove_endpoint(endpoint)
                n_endpoint = get_endpoint()
                return process_search_row(search_string,n_endpoint,entry_id)  # Add return here
            else:
                response_json = response.json()
                result = response_json.get('body', None)
                if result:
                    unpacked_html = unpack_content(result)
                    print(len(unpacked_html))
                    parsed_data = GP(unpacked_html)
                    if parsed_data is None:
                        print('trying again 2')
                        remove_endpoint(endpoint)
                        n_endpoint = get_endpoint()
                        return process_search_row(search_string,n_endpoint,entry_id)  # Add return here
                    if type(parsed_data)==list:
                        if parsed_data[0][0] == 'No start_tag or end_tag':
                            print('trying again 3')
                            remove_endpoint(endpoint)
                            n_endpoint = get_endpoint()
                            return process_search_row(search_string,n_endpoint,entry_id)
                    else:
                        print('parsed data!')
                        image_url = parsed_data[0]
                        image_desc = parsed_data[1]
                        image_source = parsed_data[2]
                        image_thumb = parsed_data[3]

                        print(
                            f'Image URL: {type(image_url)} {image_url}\nImage Desc:  {type(image_desc)} {image_desc}\nImage Source:{type(image_source)}  {image_source}')
                        if image_url:
                            df = pd.DataFrame({
                                'ImageUrl': image_url,
                                'ImageDesc': image_desc,
                                'ImageSource': image_source,
                                'ImageUrlThumbnail':image_thumb,
                            })
                            if not df.empty:
                                    df.insert(0, 'EntryId', entry_id)
                                    df.to_sql(name='utb_ImageScraperResult', con=engine, index=False,
                                                     if_exists='append')

                                    sql_query = f"update utb_ImageScraperRecords set  Step1 = getdate() where EntryID = {entry_id}"

                                    # Create a cursor from the connection
                                    connection = pyodbc.connect(conn)
                                    cursor = connection.cursor()

                                    # Execute the update query
                                    cursor.execute(sql_query)

                                    # Commit the changes
                                    connection.commit()

                                    # Close the connection
                                    connection.close()
                        else:
                            print('trying again 4')

                            remove_endpoint(endpoint)
                            n_endpoint = get_endpoint()
                            return process_search_row(search_string,n_endpoint,entry_id)
        except requests.RequestException as e:
            print('trying again 5')
            remove_endpoint(endpoint)
            n_endpoint = get_endpoint()
            print(f"Error making request: {e}\nTrying Again: {n_endpoint}")
            return process_search_row(search_string,n_endpoint,entry_id)
        
        
def update_file_generate_complete(file_id):
    query = f'update utb_ImageScraperFiles set CreateFileCompleteTime = getdate() Where ID = {file_id}'       
    connection = pyodbc.connect(conn)
    cursor = connection.cursor()

    # Execute the update query
    cursor.execute(query)
    # Commit the changes
    connection.commit()

    # Close the connection
    connection.close()    
        
def get_file_location(file_id):
    query = f"Select FileLocationUrl from utb_ImageScraperFiles where ID = {file_id}"
    connection = pyodbc.connect(conn)
    cursor = connection.cursor()

    # Execute the update query
    cursor.execute(query)
    file_location_url = cursor.fetchone()
    # Commit the changes
    connection.commit()

    # Close the connection
    connection.close()
    if file_location_url:
        (file_location_url,) = file_location_url
        print(file_location_url)
    else:

        file_location_url = "No File Found"
    return file_location_url

def update_file_location_complete(file_id,file_location):
    query = f"update utb_ImageScraperFiles set FileLocationURLComplete = '{file_location}' Where ID ={file_id}"
    connection = pyodbc.connect(conn)
    cursor = connection.cursor()

    # Execute the update query
    cursor.execute(query)

    # Commit the changes
    connection.commit()
    connection.close()
    
def get_images_excel_db(file_id):
    update_file_start_query = f"update utb_ImageScraperFiles set CreateFileStartTime = getdate() Where ID = {file_id}"
    connection = pyodbc.connect(conn)
    cursor = connection.cursor()

    # Execute the update query
    cursor.execute(update_file_start_query)

    # Commit the changes
    connection.commit()
    connection.close()

    query_get_images_to_excel = """Select s.ExcelRowID, r.ImageUrl, r.ImageUrlThumbnail from utb_ImageScraperFiles f
inner join utb_ImageScraperRecords s on s.FileID = f.ID 
inner join utb_ImageScraperResult r on r.EntryID = s.EntryID 
Where f.ID = $FileID$ and r.SortOrder = 1
Order by s.ExcelRowID"""

    query_get_images_to_excel = query_get_images_to_excel.replace('$FileID$',str(file_id))
    print(query_get_images_to_excel)
    # Close the connection

    df = pd.read_sql_query(query_get_images_to_excel, con=engine)
    return df

def update_sort_order(file_id):
    print('executing update sort order')
    query = """with toupdate as (
select t.*,
row_number() over (partition by t.EntryID order by t.ResultID) as seqnum
from utb_ImageScraperResult t 
inner join utb_ImageScraperRecords r on r.EntryID = t.EntryID 
Where r.FileID = $FileID$ ) update toupdate set SortOrder = seqnum;"""

    query = query.replace('$FileID$',str(file_id))
    connection = pyodbc.connect(conn)
    cursor = connection.cursor()

    # Execute the update queryf
    cursor.execute(query)

    # Commit the changes
    connection.commit()

    # Close the connection
    connection.close()
    
    
    
    
    
    
    
    complete_query = f"update utb_ImageScraperFiles set ImageCompleteTime = getdate() Where ID = {file_id}"
    connection = pyodbc.connect(conn)
    cursor = connection.cursor()

    # Execute the update query
    cursor.execute(complete_query)

    # Commit the changes
    connection.commit()

    # Close the connection
    connection.close()
    print('completed update sort order')
async def generate_download_file(file_id):
    preferred_image_method = 'append'
    
    start_time = time.time()
    loop = asyncio.get_running_loop()
    
    selected_images_df = await loop.run_in_executor(ThreadPoolExecutor(), get_images_excel_db, file_id)
    selected_image_list = await loop.run_in_executor(ThreadPoolExecutor(), prepare_images_for_download_dataframe,selected_images_df )
    
    print(selected_images_df.head())
    print(selected_image_list)
    
    provided_file_path = await loop.run_in_executor(ThreadPoolExecutor(), get_file_location,file_id )
    file_name = provided_file_path.split('/')[-1]
    temp_images_dir, temp_excel_dir = await create_temp_dirs(file_id)
    local_filename = os.path.join(temp_excel_dir, file_name)
    failed_img_urls = await download_all_images(selected_image_list, temp_images_dir)
    contenttype = os.path.splitext(local_filename)[1]
    response = await loop.run_in_executor(None, requests.get, provided_file_path, {'allow_redirects': True, 'timeout': 60})
    if response.status_code != 200:
        logger.error(f"Failed to download file: {response.status_code}")
        return {"error": "Failed to download the provided file."}
    with open(local_filename, "wb") as file:
        file.write(response.content)
        
    logger.info("Writing images to Excel")
    failed_rows = await loop.run_in_executor(ThreadPoolExecutor(), write_excel_image, local_filename, temp_images_dir, preferred_image_method)
    print(f"failed rows: {failed_rows}")
    #if failed_rows != []:
    if failed_img_urls:
        print(failed_img_urls)
        #fail_rows_written = await loop.run_in_executor(ThreadPoolExecutor(), write_failed_img_urls, local_filename, clean_results,failed_rows)
        fail_rows_written = await loop.run_in_executor(ThreadPoolExecutor(), write_failed_downloads_to_excel, failed_img_urls,local_filename)
        logger.error(f"Failed to write images for rows: {failed_rows}")
        logger.error(f"Failed rows added to excel: {fail_rows_written})")
        
    logger.info("Uploading file to space")
    #public_url = upload_file_to_space(local_filename, local_filename, is_public=True)
    is_public = True
    public_url = await loop.run_in_executor(ThreadPoolExecutor(), upload_file_to_space, local_filename, local_filename,is_public,contenttype)  
    await loop.run_in_executor(ThreadPoolExecutor(), update_file_location_complete, file_id, public_url)
    await loop.run_in_executor(ThreadPoolExecutor(), update_file_generate_complete, file_id)
    
    end_time = time.time()
    execution_time = end_time - start_time
    #await loop.run_in_executor(ThreadPoolExecutor(), send_email, send_to_email, 'Your File Is Ready', public_url, local_filename)
    if os.listdir(temp_images_dir) !=[]:
        logger.info("Sending email")
        #await loop.run_in_executor(ThreadPoolExecutor(), send_email, send_to_email, f'Started {file_name}', public_url, local_filename,execution_time,'')
    #await send_email(send_to_email, 'Your File Is Ready', public_url, local_filename)
    logger.info("Cleaning up temporary directories")
    await cleanup_temp_dirs([temp_images_dir, temp_excel_dir])
    
    logger.info("Processing completed successfully")
    
    return {"message": "Processing completed successfully.", "results": 'hardcoded result value', "public_url": public_url}
    
# async def process_image_batch(payload: dict):
#     start_time = time.time()
#     # Your existing logic here
#     # Include all steps from processing start to finish,
#     # such as downloading images, writing images to Excel, etc.
#
#     logger.info(f"Processing started for payload: {payload}")
#     rows = payload.get('rowData', [])
#     provided_file_path = payload.get('filePath')
#     logger.info("Received request to process image batch")
#     file_name = provided_file_path.split('/')[-1]
#     send_to_email = payload.get('sendToEmail', 'nik@iconluxurygroup.com')
#     preferred_image_method = payload.get('preferredImageMethod', 'append')
#     file_id_db = insert_file_db(file_name, provided_file_path)
#     print(file_id_db)
#     load_payload_db(rows, file_id_db)
#     search_df = get_records_to_search(file_id_db, engine)
#     print(search_df)
#
#     semaphore = asyncio.Semaphore(int(os.environ.get('MAX_THREAD')))  # Limit concurrent tasks to avoid overloading
#     loop = asyncio.get_running_loop()
#     print(rows)
#     try:
#     #    # Create a temporary directory to save downloaded images
#     #
#     #
#          await loop.run_in_executor(ThreadPoolExecutor(), send_message_email, send_to_email, f'Started {file_name}', f'Total Rows: {len(rows)}\nFilename: {file_name}\nDB_file_id: {file_id_db}\nUploaded File: {provided_file_path}')
#     #
#          tasks = [process_with_semaphore(row, semaphore,file_id_db) for _, row in search_df.iterrows()]
#          await asyncio.gather(*tasks, return_exceptions=True)
#
#          await loop.run_in_executor(ThreadPoolExecutor(), update_sort_order,file_id_db)
#     #
#          #if any(isinstance(result, Exception) for result in results):
#              #logger.error("Error occurred during image processing.")
#              #return {"error": "An error occurred during processing."}
#          #print(results)
#
#          # logger.info("Downloading images")
#          #
#          # #clean_results = await loop.run_in_executor(ThreadPoolExecutor(), prepare_images_for_downloadV2, results,send_to_email)
#          # clean_results = await loop.run_in_executor(ThreadPoolExecutor(), prepare_images_for_downloadV2, results)
#          # if clean_results == []:
#          #     send_message_email(send_to_email,f'Started {file_name}','No images found\nPlease make sure correct column values are provided\nIf api is disabled this reponse will be sent')
#          # print(clean_results)
#          # logger.info("clean_results: {}".format(clean_results))
#
#
#
#     # #
#     # #     #d_complete_ = await loop.run_in_executor(ThreadPoolExecutor(), download_all_images, clean_results, temp_images_dir)
#     #      failed_img_urls = await download_all_images(clean_results, temp_images_dir)
#     # #
#     #      contenttype = os.path.splitext(local_filename)[1]
#     #      logger.info("Downloading Excel from web")
#     #      response = await loop.run_in_executor(None, requests.get, provided_file_path, {'allow_redirects': True, 'timeout': 60})
#     #      if response.status_code != 200:
#     #          logger.error(f"Failed to download file: {response.status_code}")
#     #          return {"error": "Failed to download the provided file."}
#     #      with open(local_filename, "wb") as file:
#     #          file.write(response.content)
#
#     #     logger.info("Writing images to Excel")
#     #     failed_rows = await loop.run_in_executor(ThreadPoolExecutor(), write_excel_image, local_filename, temp_images_dir, preferred_image_method)
#     #     print(f"failed rows: {failed_rows}")
#     #     #if failed_rows != []:
#     #     if failed_img_urls:
#     #         print(failed_img_urls)
#     #         #fail_rows_written = await loop.run_in_executor(ThreadPoolExecutor(), write_failed_img_urls, local_filename, clean_results,failed_rows)
#     #         fail_rows_written = await loop.run_in_executor(ThreadPoolExecutor(), write_failed_downloads_to_excel, failed_img_urls,local_filename)
#     #         logger.error(f"Failed to write images for rows: {failed_rows}")
#     #         logger.error(f"Failed rows added to excel: {fail_rows_written})")
#     #
#     #     logger.info("Uploading file to space")
#     #     #public_url = upload_file_to_space(local_filename, local_filename, is_public=True)
#     #     is_public = True
#     #     public_url = await loop.run_in_executor(ThreadPoolExecutor(), upload_file_to_space, local_filename, local_filename,is_public,contenttype)
#     #     end_time = time.time()
#     #     execution_time = end_time - start_time
#     #     #await loop.run_in_executor(ThreadPoolExecutor(), send_email, send_to_email, 'Your File Is Ready', public_url, local_filename)
#     #     if os.listdir(temp_images_dir) !=[]:
#     #         logger.info("Sending email")
#     #         await loop.run_in_executor(ThreadPoolExecutor(), send_email, send_to_email, f'Started {file_name}', public_url, local_filename,execution_time,'')
#     #     #await send_email(send_to_email, 'Your File Is Ready', public_url, local_filename)
#     #     logger.info("Cleaning up temporary directories")
#     #     await cleanup_temp_dirs([temp_images_dir, temp_excel_dir])
#     #
#     #     logger.info("Processing completed successfully")
#     #
#     #     return {"message": "Processing completed successfully.", "results": results, "public_url": public_url}
#     #
#     except Exception as e:
#          logger.exception("An unexpected error occurred during processing: %s", e)
#          await loop.run_in_executor(ThreadPoolExecutor(), send_message_email, send_to_email, f'Started {file_name}', f"An unexpected error occurred during processing.\nError: {str(e)}")
#          return {"error": f"An unexpected error occurred during processing. Error: {e}"}
#
import asyncio
import ray

def get_lm_products(file_id):

    connection = pyodbc.connect(conn)
    cursor = connection.cursor()
    query = f"exec usp_ImageScrapergetMatchFromRetail {file_id}"
    print(query)
    # Execute the update query
    cursor.execute(query)

    # Commit the changes
    connection.commit()
    connection.close()

def process_image_batch(payload: dict):
    logger.info(f"Processing started for payload: {payload}")
    rows = payload.get('rowData', [])
    provided_file_path = payload.get('filePath')
    logger.info("Received request to process image batch")
    file_name = provided_file_path.split('/')[-1]
    send_to_email = payload.get('sendToEmail', 'nik@iconluxurygroup.com')
    preferred_image_method = payload.get('preferredImageMethod', 'append')
    file_id_db = insert_file_db(file_name, provided_file_path)
    print(file_id_db)
    load_payload_db(rows, file_id_db)
    get_lm_products(file_id_db)
    search_df = get_records_to_search(file_id_db, engine)
    print(search_df)
    search_list=list(search_df.T.to_dict().values())
    ####
    start = datetime.datetime.now()
    print(f"Start of whole process: {start}")
    BATCH_SIZE=100
    batches=[search_list[i:i+BATCH_SIZE] for i in range(0, len(search_list), BATCH_SIZE)]
    print(f"Batches: {batches} CCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCCC")
    futures=[process_batch.remote(batch) for batch in batches]
    ray.get(futures)
    end = datetime.datetime.now()
    print(f"End of whole process: {end}")
    print(f"It took {end - start} to complete")
    #####
    update_sort_order(file_id_db)


@app.post("/process-image-batch/")
async def process_payload(background_tasks: BackgroundTasks, payload: dict):
    logger.info("Received request to process image batch")
    background_tasks.add_task(process_image_batch, payload)
    #await process_image_batch(payload)
    return {"message": "Processing started successfully. You will be notified upon completion."}
@app.post("/generate-download-file/")
async def process_file(background_tasks: BackgroundTasks, file_id: int):
    logger.info("Received request to process image batch")
    background_tasks.add_task(generate_download_file, str(file_id))
    #await generate_download_file((str(file_id)))
    return {"message": "Processing started successfully. You will be notified upon completion."}

@ray.remote
def process_db_row(row):
    entry_id = row['EntryID']
    searchString = row['SearchString']
    print(f"Entry Id: {entry_id}\nSearch String {searchString}")
    endpoint = get_endpoint()
    process_search_row(searchString, endpoint, entry_id)
@ray.remote
def process_batch(batch):
    # Process each item in the batch in parallel
    futures = [process_db_row.remote(data) for data in batch]
    print(futures)
    results = ray.get(futures)
    return results



def highlight_cell(excel_file, cell_reference):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    sheet[cell_reference].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    workbook.save(excel_file)
def write_failed_downloads_to_excel(failed_downloads, excel_file):
    if failed_downloads:
        workbook = openpyxl.load_workbook(excel_file)

        # Select the active worksheet or specify by name
        worksheet = workbook.active  # or workbook.get_sheet_by_name('SheetName')
        # Iterate over the failed rows
        for row in failed_downloads:
            url = row[0]
            row_id = row[1]
            if url:
                if url != 'None found in this filter':
                    # Write the URL to column A of the failed row
                    # Adjust the cell reference as needed (row index might need +1 depending on header row)
                    cell_reference = f"{get_column_letter(1)}{row_id}"  # Column A, row number
                    worksheet[cell_reference] = str(url)
                    highlight_cell(excel_file, cell_reference)
        workbook.save(excel_file)
        logger.info(f"Failed downloads written to Excel file: {excel_file}")
    else:
        logger.info("No failed downloads to write to Excel.")
def write_failed_img_urls(excel_file_path, clean_results, failed_rows):
    # Load the workbook
    added_rows = [] 
    workbook = openpyxl.load_workbook(excel_file_path)
    
    # Select the active worksheet or specify by name
    worksheet = workbook.active  # or workbook.get_sheet_by_name('SheetName')
    
    # Convert clean_results to a dictionary for easier lookup
    clean_results_dict = {row: url for row, url in clean_results}
    
    # Iterate over the failed rows
    for row in failed_rows:
        # Look up the URL in the clean_results_dict using the row as a key
        url = clean_results_dict.get(row)
        
        if url:
            # Write the URL to column A of the failed row
            # Adjust the cell reference as needed (row index might need +1 depending on header row)
            cell_reference = f"{get_column_letter(1)}{row}"  # Column A, row number
            worksheet[cell_reference] = str(url)
            highlight_cell(excel_file_path,cell_reference)
            added_rows.append(row)
            
        # Save the workbook
        workbook.save(excel_file_path)
    return added_rows
def prepare_images_for_download_dataframe(df):
    images_to_download = []

    for row in df.itertuples(index=False,name=None):
        print(row)
        if row[1] != 'No google image results found':
            images_to_download.append(row)

    return images_to_download

def prepare_images_for_download(results,send_to_email):
    images_to_download = []

    for package in results:
        # Ensure the 'result' key is available and its 'status' is 'Completed'.
        if package.get('result', {}).get('url') == 'Completed':
            # Iterate over each 'result' entry if it exists and is a list.
            for result_entry in package.get('result', {}).get('result', []):
                # Check if the entry is 'Completed' and contains a 'result' key with a URL.
                if result_entry.get('status') == 'Completed' and isinstance(result_entry.get('result'), dict):
                    result_data = result_entry.get('result')
                    url = result_data.get('url')
                    if url:  # Ensure the URL is not None or empty.
                        images_to_download.append((package.get('absoluteRowIndex'), url))

    #if not images_to_download:
        #send_message_email(send_to_email,f'Started {file_name}','No images found in the results')
        #raise Exception("No valid image URLs found in the results")


    return images_to_download
import tldextract
from collections import Counter
def extract_domains_and_counts(data):
    """Extract domains from URLs and count their occurrences."""
    domains = [tldextract.extract(url).registered_domain for _, url,thumb in data]
    domain_counts = Counter(domains)
    return domain_counts

def analyze_data(data):
    domain_counts = extract_domains_and_counts(data)
    logger.info("Domain counts: %s", domain_counts)
    unique_domains = len(domain_counts)
    print(f"Unique Domain Len: {unique_domains}")
    pool_size = min(500, max(10, unique_domains * 2))  # Adjust as needed
    print(f"Pool size: {pool_size}")
    return pool_size
def build_headers(url):
    domain_info = tldextract.extract(url)
    domain = f"{domain_info.domain}.{domain_info.suffix}"
    
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "en-US,en;q=0.9",
        "Upgrade-Insecure-Requests": "1",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        # "Referer": "Set this if needed based on your logic"
    }
    #if domain:
        #headers["Referer"] = f"https://{domain}/"
        #print(f"Headers: {headers['Referer']}")
    # Additional dynamic header settings can go here.
    # Example for Referer (if applicable):
    # headers["Referer"] = f"https://{domain}/"
    
    return headers
def try_convert_to_png(image_path, new_path, image_name):
    logger.info(f"Attempting to convert image to PNG: {image_path}")
    try:
        with IMG2.open(image_path) as img:
            final_image_path = os.path.join(new_path, f"{image_name}.png")
            img.convert("RGB").save(final_image_path, 'PNG')
            os.remove(image_path)  # Cleanup original/temp file
            logger.info(f"Image successfully converted to PNG: {final_image_path}")
            return True
    except IOError as e:
        logger.error(f"Failed to convert image to PNG: {e}")
        return False
async def download_all_images(data, save_path):
    failed_downloads = []
    pool_size = analyze_data(data)  # Placeholder for your actual data analysis function

    logger.info(f"Setting up session with pool size: {pool_size}")

    # Setup async session with retry policy
    timeout = ClientTimeout(total=60)
    retry_options = ExponentialRetry(attempts=3, start_timeout=3)
    connector = aiohttp.TCPConnector(ssl=False)

    async with RetryClient(raise_for_status=False, retry_options=retry_options, timeout=timeout, connector=connector) as session:
        semaphore = asyncio.Semaphore(pool_size)

        logger.info("Scheduling image downloads")
        tasks = [
            image_download(semaphore, str(item[1]),str(item[2]), str(item[0]), save_path, session, index)
            for index, item in enumerate(data, start=1)
        ]

        results = await asyncio.gather(*tasks, return_exceptions=True)

        logger.info("Processing download results")
        for index, result in enumerate(results):
            if isinstance(result, Exception):
                #THUMBNAIL DOWNLOAD ON FAIL
                print(data[index])
                logger.error(f"Download task generated an exception: {result}")
                logger.error(f"Trying again with :{str(data[index][2])}")
                print(f"Download task generated an exception: {result}")
                print(f"Trying again with :{str(data[index][2])}")
                await thumbnail_download(semaphore, str(data[index][2]),str(data[index][0]), save_path, session, fallback_formats=None)
                #THUMBNAIL DOWNLOAD ON FAIL
                failed_downloads.append((data[index][1], data[index][0]))  # Append the image URL and row ID
            else:
                logger.info(f"Download task completed with result: {result}")
                if result is False:
                    failed_downloads.append((data[index][1], data[index][0]))  # Append the image URL and row ID

    return failed_downloads


async def image_download(semaphore, url, thumbnail, image_name, save_path, session, fallback_formats=None):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "Accept-Language": "en-US,en;q=0.9",
    }
    async with semaphore:
        if fallback_formats is None:
            fallback_formats = ['png', 'jpeg', 'gif', 'bmp', 'webp', 'avif', 'tiff', 'ico']

        logger.info(f"Initiating download for URL: {url} Img: {image_name}")
        try:
            async with session.get(url, headers=headers) as response:
                logger.info(f"Requesting URL: {url} with stream=True")
                # response = session.get(url, stream=True)
                logger.info(f"Received response: {response.status} for URL: {url}")

                if response.status == 200:
                    logger.info(f"Processing content from URL: {url}")
                    data = await response.read()
                    image_data = BytesIO(data)
                    try:
                        logger.info(f"Attempting to open image stream and save as PNG for {image_name}")
                        with IMG2.open(image_data) as img:
                            final_image_path = os.path.join(save_path, f"{image_name}.png")
                            img.save(final_image_path)
                            logger.info(f"Successfully saved: {final_image_path}")
                            return True
                    except UnidentifiedImageError as e:
                        logger.error(f"Image file type unidentified, trying fallback formats for {image_name}: {e}")
                        for fmt in fallback_formats:
                            image_data.seek(0)  # Reset stream position
                            try:
                                logger.info(f"Trying to save image with fallback format {fmt} for {image_name}")
                                with IMG2.open(image_data) as img:
                                    final_image_path = os.path.join(save_path, f"{image_name}.{fmt}")
                                    img.save(final_image_path)
                                    logger.info(f"Successfully saved with fallback format {fmt}: {final_image_path}")
                                    return True
                            except Exception as fallback_exc:
                                logger.error(f"Failed with fallback format {fmt} for {image_name}: {fallback_exc}")
                else:
                    logger.error(f"Download failed with status code {response.status} for URL: {url}")
                    await thumbnail_download(semaphore, thumbnail, image_name, save_path, session,
                                             fallback_formats=None)

        except TimeoutError as exc:
            # Handle the timeout specifically
            logger.error(f"Timeout occurred while downloading {url} Image: {image_name}")
            print('timeout error inside the downlaod function')
            # await thumbnail_download(semaphore, thumbnail ,image_name, save_path, session, fallback_formats=None)
            # return False
            return exc
        except Exception as exc:
            logger.error(f"Exception occurred during download or processing for URL: {url}: {exc}", exc_info=True)
            # await thumbnail_download(semaphore, thumbnail ,image_name, save_path, session, fallback_formats=None)
            # return False
            return exc
async def thumbnail_download(semaphore, url, image_name, save_path, session, fallback_formats=None):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "Accept-Language": "en-US,en;q=0.9",
    }
    async with semaphore:
        if fallback_formats is None:
            fallback_formats = ['png', 'jpeg', 'gif', 'bmp', 'webp', 'avif', 'tiff', 'ico']

        logger.info(f"Initiating download for URL: {url} Img: {image_name}")
        try:
            async with session.get(url, headers=headers) as response:
                logger.info(f"Requesting URL: {url} with stream=True")
                # response = session.get(url, stream=True)
                logger.info(f"Received response: {response.status} for URL: {url}")

                if response.status == 200:
                    logger.info(f"Processing content from URL: {url}")
                    data = await response.read()
                    image_data = BytesIO(data)
                    try:
                        logger.info(f"Attempting to open image stream and save as PNG for {image_name}")
                        with IMG2.open(image_data) as img:
                            final_image_path = os.path.join(save_path, f"{image_name}.png")
                            img.save(final_image_path)
                            logger.info(f"Successfully saved: {final_image_path}")
                            return True
                    except UnidentifiedImageError as e:
                        logger.error(f"Image file type unidentified, trying fallback formats for {image_name}: {e}")
                        for fmt in fallback_formats:
                            image_data.seek(0)  # Reset stream position
                            try:
                                logger.info(f"Trying to save image with fallback format {fmt} for {image_name}")
                                with IMG2.open(image_data) as img:
                                    final_image_path = os.path.join(save_path, f"{image_name}.{fmt}")
                                    img.save(final_image_path)
                                    logger.info(f"Successfully saved with fallback format {fmt}: {final_image_path}")
                                    return True
                            except Exception as fallback_exc:
                                logger.error(f"Failed with fallback format {fmt} for {image_name}: {fallback_exc}")
                else:
                    logger.error(f"Download failed with status code {response.status} for URL: {url}")

        except TimeoutError:
            # Handle the timeout specifically
            logger.error(f"Timeout occurred while downloading {url} Image: {image_name}")
            return False
        except Exception as exc:
            logger.error(f"Exception occurred during download or processing for URL: {url}: {exc}", exc_info=True)
            return False
def verify_png_image_single(image_path):
    try:
        img = IMG2.open(image_path)
        img.verify()  # I'm using verify() method to check if it's a valid image
        logging.info(f"Image verified successfully: {image_path}")
    except Exception as e:
        logging.error(f"IMAGE verify ERROR: {e}, for image: {image_path}")
        return False

    imageSize = os.path.getsize(image_path)
    logging.debug(f"Image size: {imageSize} bytes")

    if imageSize < 3000:
        logging.warning(f"File may be corrupted or too small: {image_path}")
        return False

    try:
        resize_image(image_path)
    except Exception as e:
        logging.error(f"Error resizing image: {e}, for image: {image_path}")
        return False
    return True

def resize_image(image_path):
    try:
        img = IMG2.open(image_path)
        MAXSIZE = 145
        if img:
            h, w = img.height, img.width  # original size
            logging.debug(f"Original size: height={h}, width={w}")
            if h > MAXSIZE or w > MAXSIZE:
                if h > w:
                    w = int(w * MAXSIZE / h)
                    h = MAXSIZE
                else:
                    h = int(h * MAXSIZE / w)
                    w = MAXSIZE
            logging.debug(f"Resized to: height={h}, width={w}")
            newImg = img.resize((w, h))
            newImg.save(image_path)
            logging.info(f"Image resized and saved: {image_path}")
            return True
    except Exception as e:
        logging.error(f"Error resizing image: {e}, for image: {image_path}")
        return False               
def write_excel_image(local_filename, temp_dir,preferred_image_method):
    failed_rows = []
    # Load the workbook and select the active worksheet
    wb = load_workbook(local_filename)
    ws = wb.active
    print(os.listdir(temp_dir))
    
    # Iterate through each file in the temporary directory
    for image_file in os.listdir(temp_dir):
        image_path = os.path.join(temp_dir, image_file)
        # Extract row number or other identifier from the image file name
        try:
            # Assuming the file name can be directly converted to an integer row number
            row_number = int(image_file.split('.')[0])
            logging.info(f"Processing row {row_number}, image path: {image_path}")
        except ValueError:
            logging.warning(f"Skipping file {image_file}: does not match expected naming convention")
            continue  # Skip files that do not match the expected naming convention
        verify_image = verify_png_image_single(image_path)    
        # Check if the image meets the criteria to be added
        if verify_image:
            logging.info('Inserting image')
            img = openpyxl.drawing.image.Image(image_path)
            # Determine the anchor point based on the preferred image method
            if preferred_image_method in ["overwrite", "append"]:
                anchor = "A" + str(row_number)
                logging.info('Anchor assigned')
            elif preferred_image_method == "NewColumn":
                anchor = "B" + str(row_number)  # Example adjustment for a different method
            else:
                logging.error(f'Unrecognized preferred image method: {preferred_image_method}')
                continue  # Skip if the method is not recognized
                
            img.anchor = anchor
            ws.add_image(img)
            #wb.save(local_filename)
            logging.info(f'Image saved at {anchor}')
        else:
            failed_rows.append(row_number)
            logging.warning('Inserting image skipped due to verify_png_image_single failure.')   
    # Finalize changes to the workbook
    logging.info('Finished processing all images.')
    wb.save(local_filename)
    return failed_rows 


if __name__ == "__main__":
    logger.info("Starting Uvicorn server")
    print(os.environ)
    #uvicorn.run("main:app", port=8080, host='0.0.0.0', reload=True)
    uvicorn.run("main:app", port=8080, host='0.0.0.0')
    if ray.is_initialized():
        ray.shutdown()
    ray.init(address='auto')

